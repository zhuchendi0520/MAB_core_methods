#!/usr/bin/env python3
"""
Step 02: add exact-binomial p-values and BH-FDR columns to pN/pS blocks.

Expected pNS block:
  pNS, total_events, observed_syn, observed_nonsyn, expected_syn, expected_nonsyn

For each block, the script uses a two-sided exact binomial test:
  observed_nonsyn ~ Binomial(total_events, expected_nonsyn / (expected_syn + expected_nonsyn))

BH-FDR is calculated separately for each pNS block.
"""

from __future__ import annotations

import math
import sys
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def usage() -> None:
    print("Usage: python add_pns_fdr_to_wide_xlsx.py input.xlsx [output.xlsx]")


if len(sys.argv) not in (2, 3):
    usage()
    sys.exit(1)

input_xlsx = Path(sys.argv[1])
output_xlsx = (
    Path(sys.argv[2])
    if len(sys.argv) == 3
    else input_xlsx.with_name(input_xlsx.stem + "_with_fdr.xlsx")
)


def as_float(value):
    if value is None or value == "" or value == "-":
        return None
    if isinstance(value, str) and value.strip().upper() in {"NA", "NAN"}:
        return None
    try:
        return float(value)
    except Exception:
        return None


def log_comb(n: int, k: int) -> float:
    if k < 0 or k > n:
        return float("-inf")
    return math.lgamma(n + 1) - math.lgamma(k + 1) - math.lgamma(n - k + 1)


def binom_prob(n: int, k: int, p: float) -> float:
    if k < 0 or k > n:
        return 0.0
    if p <= 0:
        return 1.0 if k == 0 else 0.0
    if p >= 1:
        return 1.0 if k == n else 0.0
    return math.exp(log_comb(n, k) + k * math.log(p) + (n - k) * math.log1p(-p))


def binom_two_sided_pvalue(k: int, n: int, p: float):
    if n <= 0 or p < 0 or p > 1:
        return None
    observed = binom_prob(n, k, p)
    total = 0.0
    eps = 1e-12
    for x in range(n + 1):
        px = binom_prob(n, x, p)
        if px <= observed + eps:
            total += px
    return min(1.0, total)


def bh_fdr(p_values):
    indexed = [(i, p) for i, p in enumerate(p_values) if p is not None]
    out = [None] * len(p_values)
    if not indexed:
        return out

    indexed.sort(key=lambda x: x[1])
    m = len(indexed)
    running_min = 1.0
    for rank_from_end, (idx, p) in enumerate(reversed(indexed), start=1):
        rank = m - rank_from_end + 1
        q = min(running_min, p * m / rank)
        out[idx] = min(1.0, q)
        running_min = q
    return out


def copy_cell(src, dst, value):
    dst.value = value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


in_wb = load_workbook(input_xlsx)
in_ws = in_wb.active
value_wb = load_workbook(input_xlsx, data_only=True)
value_ws = value_wb.active
headers = [in_ws.cell(1, c).value for c in range(1, in_ws.max_column + 1)]

pns_cols = [i + 1 for i, h in enumerate(headers) if h == "pNS"]
if not pns_cols:
    raise SystemExit("No pNS columns found.")

expected_following = [
    "total_events",
    "observed_syn",
    "observed_nonsyn",
    "expected_syn",
    "expected_nonsyn",
]

blocks = []
for pns_col in pns_cols:
    following = [in_ws.cell(1, pns_col + i).value for i in range(1, 6)]
    if following != expected_following:
        raise SystemExit(
            f"pNS column {pns_col} is not followed by expected columns: {following}"
        )

    p_values = []
    for row in range(2, in_ws.max_row + 1):
        total = as_float(value_ws.cell(row, pns_col + 1).value)
        observed_nonsyn = as_float(value_ws.cell(row, pns_col + 3).value)
        expected_syn = as_float(value_ws.cell(row, pns_col + 4).value)
        expected_nonsyn = as_float(value_ws.cell(row, pns_col + 5).value)

        if None in (total, observed_nonsyn, expected_syn, expected_nonsyn):
            p_values.append(None)
            continue

        n = int(round(total))
        k = int(round(observed_nonsyn))
        denom = expected_syn + expected_nonsyn
        if n <= 0 or denom <= 0:
            p_values.append(None)
            continue

        p_expected_nonsyn = expected_nonsyn / denom
        p_values.append(binom_two_sided_pvalue(k, n, p_expected_nonsyn))

    blocks.append(
        {
            "pns_col": pns_col,
            "p_values": p_values,
            "fdr_values": bh_fdr(p_values),
        }
    )

block_by_pns_col = {b["pns_col"]: b for b in blocks}

out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = in_ws.title

header_fill = PatternFill("solid", fgColor="D9EAD3")
header_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")

out_col = 1
for in_col in range(1, in_ws.max_column + 1):
    for row in range(1, in_ws.max_row + 1):
        copy_cell(
            in_ws.cell(row, in_col),
            out_ws.cell(row, out_col),
            value_ws.cell(row, in_col).value,
        )

    if in_col in block_by_pns_col:
        block = block_by_pns_col[in_col]
        for label, values in [
            ("pNS_p_value", block["p_values"]),
            ("pNS_FDR", block["fdr_values"]),
        ]:
            out_col += 1
            h = out_ws.cell(1, out_col)
            h.value = label
            h.font = header_font
            h.fill = header_fill
            h.alignment = center
            for row_idx, value in enumerate(values, start=2):
                cell = out_ws.cell(row_idx, out_col)
                cell.value = value
                cell.number_format = "0.0000"

    out_col += 1

for col_idx in range(1, out_ws.max_column + 1):
    col_letter = out_ws.cell(1, col_idx).column_letter
    header = out_ws.cell(1, col_idx).value
    if header in {"pNS_p_value", "pNS_FDR"}:
        out_ws.column_dimensions[col_letter].width = 14
    else:
        out_ws.column_dimensions[col_letter].width = in_ws.column_dimensions[
            in_ws.cell(1, min(col_idx, in_ws.max_column)).column_letter
        ].width or 12

out_ws.freeze_panes = "A2"
out_wb.save(output_xlsx)

for i, block in enumerate(blocks, start=1):
    tested = sum(p is not None for p in block["p_values"])
    sig = sum(q is not None and q < 0.05 for q in block["fdr_values"])
    print(f"pNS block {i}: tested={tested}, FDR<0.05={sig}")
print(output_xlsx)
